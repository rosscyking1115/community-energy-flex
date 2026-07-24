"""Write an Excel action workbook. Requires the ``reports`` extra (openpyxl)."""

from __future__ import annotations

from io import BytesIO

from community_energy_flex.reporting.summary import (
    ActionSummary,
    ReportingContext,
    _report_language,
)


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:  # pragma: no cover - depends on env
        raise ImportError(
            "Excel reports need openpyxl. Install with: pip install '.[reports]'"
        ) from exc
    return openpyxl


def build_workbook(
    summary: ActionSummary, *, reporting_context: ReportingContext | None = None
):
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    context = reporting_context or ReportingContext(
        basis="forecast",
        status="not_reportable",
        reason=(
            "No explicit preferred start was supplied; "
            "cost and carbon differences are unavailable."
        ),
    )
    heading, metric_prefix = _report_language(context)
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"], ws["A1"].font = "Community Energy Flex - Action Report", bold
    ws["A3"] = "Objective"
    ws["B3"] = summary.objective
    ws["A2"] = heading
    ws["A4"] = f"{metric_prefix} cost difference"
    ws["B4"] = round(summary.total_cost_saving_pounds, 2) if context.is_reportable else None
    ws["A5"] = f"{metric_prefix} carbon difference"
    ws["B5"] = round(summary.total_carbon_saving_kg, 2) if context.is_reportable else None

    sched = wb.create_sheet("Tomorrow Schedule")
    headers = [
        "Device", "Recommended", "Baseline", "Saving (p)", "Saving (gCO2)",
        "Robustness indicator",
    ]
    sched.append(headers)
    for cell in sched[1]:
        cell.font = bold
    for line in summary.lines:
        sched.append([
            line.device_type,
            line.recommended_window,
            line.baseline_window if context.is_reportable else None,
            line.cost_saving_p if context.is_reportable else None,
            line.carbon_saving_g if context.is_reportable else None,
            line.robustness_band,
        ])

    caveats = wb.create_sheet("Caveats")
    caveats["A1"], caveats["A1"].font = "Assumptions & caveats", bold
    row = 3
    for line in summary.lines:
        caveats[f"A{row}"] = f"{line.device_type}: {line.caveat}"
        row += 1
    caveats[f"A{row + 1}"] = summary.safety_statement
    return wb


def write_workbook_bytes(
    summary: ActionSummary, *, reporting_context: ReportingContext | None = None
) -> bytes:
    """Serialise the workbook to bytes (for Streamlit download buttons)."""
    buffer = BytesIO()
    build_workbook(summary, reporting_context=reporting_context).save(buffer)
    return buffer.getvalue()


def write_workbook(
    summary: ActionSummary, path: str, *, reporting_context: ReportingContext | None = None
) -> str:
    build_workbook(summary, reporting_context=reporting_context).save(path)
    return path
